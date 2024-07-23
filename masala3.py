import os
import csv
from openpyxl import load_workbook

class LoadExcel:
    def __init__(self, filename):
        self.filename = filename
        self.check_file_exists()

    def check_file_exists(self):
        if not os.path.isfile(self.filename):
            raise FileNotFoundError(f"Fayl topilmadi: {self.filename}")

    def reader(self):
        workbook = load_workbook(filename=self.filename)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            print(row)

class LoadCsv:
    def __init__(self, filename):
        self.filename = filename
        self.check_file_exists()

    def check_file_exists(self):
        if not os.path.isfile(self.filename):
            raise FileNotFoundError(f"Fayl topilmadi: {self.filename}")

    def reader(self):
        with open(self.filename, mode='r', newline='', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                print(row)

csv_filename = 'user.csv'
csv_loader = LoadCsv(csv_filename)
csv_loader.reader()


excel_filename = 'user.xlsx'
excel_loader = LoadExcel(excel_filename)
excel_loader.reader()